"""
AR Build verification — runs the proposed build engine against every available
build day's source files and compares its output to the hand-built target
workbook row-by-row.

Read-only. Produces a diff report to stdout. Does not write any output files.

Usage:
    python tools/verify_ar_build.py
"""

from __future__ import annotations

import os
import re
import sys
from collections import Counter, defaultdict
from dataclasses import dataclass, field, replace
from datetime import datetime, timedelta
from typing import Any

import openpyxl

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
ASSETS = os.path.join(ROOT, "app/AR_AGING_assets")
EXTRA_TARGETS_DIR = os.path.join(ROOT, "app/assets/images")


# ---------------------------------------------------------------------------
# Build discovery
# ---------------------------------------------------------------------------

@dataclass
class BuildSpec:
    folder: str
    target_workbook: str
    target_date: datetime
    qbo_collection: str
    qbo_schedule: str
    tab_bank: str
    tms_reconcile: str


_DATE_RE = re.compile(r"AR_AGING_(\d\d)_(\d\d)_(\d{4})", re.I)


def _parse_target_date(filename: str) -> datetime | None:
    m = _DATE_RE.search(os.path.basename(filename))
    if not m:
        return None
    mm, dd, yyyy = m.group(1), m.group(2), m.group(3)
    try:
        return datetime(int(yyyy), int(mm), int(dd))
    except ValueError:
        return None


def discover_builds() -> list[BuildSpec]:
    candidate_folders = []
    # All immediate subdirs of AR_AGING_assets/
    for d in os.listdir(ASSETS):
        sub = os.path.join(ASSETS, d)
        if os.path.isdir(sub):
            candidate_folders.append(sub)
    # The weirdly-nested 5/19 folder
    nested = os.path.join(ASSETS, "5/20 data (5/5.19")
    if os.path.isdir(nested):
        candidate_folders.append(nested)

    builds: list[BuildSpec] = []

    for folder in candidate_folders:
        try:
            files = os.listdir(folder)
        except Exception:
            continue

        # Find target workbook (AR_AGING_*) — may be missing if target lives elsewhere
        targets = [f for f in files if f.startswith("AR_AGING_") and f.endswith(".xlsx")]
        if not targets:
            # No target in this folder; skip (we'll match by date below if it's a sources-only folder)
            target_path = None
            target_date = None
        else:
            target_path = os.path.join(folder, targets[0])
            target_date = _parse_target_date(targets[0])

        # Find sources
        tab = [f for f in files if f.startswith("Collection_Payment") and f.endswith(".xlsx")]
        col = [f for f in files if "Daily Collection" in f and f.endswith(".xlsx")]
        sch = [f for f in files if "Daily Schedule" in f and f.endswith(".xlsx")]
        tms = [f for f in files if f.startswith("APAR RECONCILE") and f.endswith(".xlsx")]

        if not (tab and col and sch and tms):
            continue  # Incomplete

        # If folder has no target but matches a known external one, link it
        if target_path is None:
            # Look at TAB BANK post date to infer
            wb = openpyxl.load_workbook(os.path.join(folder, tab[0]), data_only=True, read_only=True)
            ws = wb.active
            for r in ws.iter_rows(values_only=True):
                if r and r[4]:
                    try:
                        target_date = r[4] if isinstance(r[4], datetime) else datetime.strptime(str(r[4]), "%Y-%m-%d")
                        break
                    except Exception:
                        continue
            wb.close()
            if target_date:
                guess = f"AR_AGING_{target_date.month:02d}_{target_date.day:02d}_{target_date.year}.xlsx"
                guess_path = os.path.join(EXTRA_TARGETS_DIR, guess)
                if os.path.exists(guess_path):
                    target_path = guess_path

        if target_path is None or target_date is None:
            continue

        builds.append(BuildSpec(
            folder=folder,
            target_workbook=target_path,
            target_date=target_date,
            qbo_collection=os.path.join(folder, col[0]),
            qbo_schedule=os.path.join(folder, sch[0]),
            tab_bank=os.path.join(folder, tab[0]),
            tms_reconcile=os.path.join(folder, tms[0]),
        ))

    # Sort by target date
    builds.sort(key=lambda b: b.target_date)
    return builds


def find_input_workbook(target_date: datetime, all_builds: list[BuildSpec]) -> str | None:
    """Locate the most-recent-prior AR_AGING workbook from ANY known location."""
    candidates: list[tuple[datetime, str]] = []
    # Include all build targets
    for b in all_builds:
        if b.target_date < target_date:
            candidates.append((b.target_date, b.target_workbook))
    # Include workbooks in fallback dirs
    for root_dir in (EXTRA_TARGETS_DIR, ASSETS):
        if not os.path.isdir(root_dir):
            continue
        for f in os.listdir(root_dir):
            d = _parse_target_date(f)
            if d and d < target_date:
                candidates.append((d, os.path.join(root_dir, f)))
    if not candidates:
        return None
    candidates.sort(key=lambda x: x[0])
    return candidates[-1][1]


# ---------------------------------------------------------------------------
# AR row model
# ---------------------------------------------------------------------------

@dataclass
class ARRow:
    new_id: str | None = None
    company: str | None = None
    inv: str | None = None
    equipment: Any = None
    date: Any = None
    aging: int | None = None
    ref_no: Any = None
    mbl_no: Any = None
    amount: float | None = None
    paid: float | None = None
    balance: float | None = None
    memo: str | None = None
    ar_status: str | None = None
    wo: str | None = None

    @classmethod
    def from_tuple(cls, r: tuple) -> "ARRow":
        return cls(
            new_id=r[0], company=r[1], inv=r[2], equipment=r[3], date=r[4],
            aging=r[5], ref_no=r[6], mbl_no=r[7], amount=r[8], paid=r[9],
            balance=r[10], memo=r[11], ar_status=r[12], wo=r[13],
        )


def aging_bucket(days: int | None) -> str:
    if days is None:
        return ""
    if days < 30:
        return "A.0~29"
    if days < 60:
        return "B.30~59"
    if days < 90:
        return "C.60~89"
    if days < 120:
        return "D.90~119"
    return "E.120+"


# ---------------------------------------------------------------------------
# Loaders
# ---------------------------------------------------------------------------

def load_ar_sheet(workbook_path: str) -> tuple[list[ARRow], datetime | None]:
    wb = openpyxl.load_workbook(workbook_path, data_only=True)
    ar_sheet_name = None
    for sn in wb.sheetnames:
        if sn.startswith("AR_") and sn.count("_") == 3:
            ar_sheet_name = sn
            break
    if ar_sheet_name is None:
        return [], None
    ws = wb[ar_sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    sheet_date = None
    m = re.match(r"AR_(\d\d)_(\d\d)_(\d\d)", ar_sheet_name)
    if m:
        sheet_date = datetime(2000 + int(m.group(3)), int(m.group(1)), int(m.group(2)))
    out = [ARRow.from_tuple(r) for r in rows[1:] if r and any(c is not None for c in r)]
    return out, sheet_date


def load_qbo_collection(path: str) -> list[dict]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    lines = []
    for r in rows:
        if r is None or not r[2] or r[2] not in ("Invoice", "Payment"):
            continue
        lines.append({
            "payment_date": r[1], "txn_type": r[2], "txn_number": r[3],
            "customer": r[4], "invoice_or_ref": r[5], "amount": r[6],
            "open_balance": r[7], "account": r[8] if len(r) > 8 else None,
        })
    return lines


def load_qbo_schedule(path: str) -> list[dict]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    out = []
    for r in rows:
        if not r or not r[0] or not isinstance(r[1], str) or r[1] != "Invoice":
            continue
        out.append({
            "date": r[0], "txn_type": r[1], "customer": r[2], "inv": r[3],
            "ref": r[4], "cntr_chassis": r[5], "bl": r[6], "amount": r[7],
        })
    return out


def load_tab_bank(path: str) -> list[dict]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    out = []
    for r in rows[1:]:
        if not r or not r[0]:
            continue
        out.append({
            "check": r[0], "amount": r[1], "debtor_name": r[2], "debtor_code": r[3],
            "post_date": r[4], "pmt_type": r[5], "invoice": r[6],
            "invoice_date": r[7], "purchase_date": r[8],
            "invoice_amount": r[9], "collected_amount": r[10],
            "chargeback_amount": r[11], "po": r[12], "desc": r[13],
        })
    return out


def load_tms_reconcile(path: str) -> list[dict]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    header = None
    out = []
    for r in rows:
        if r and r[0] == "" and r[1] == "NO":
            header = r
            continue
        if not r or not header:
            continue
        if not isinstance(r[1], (int, float)):
            continue
        out.append({
            "no": r[1], "fold": r[2], "wo_div": r[3], "type": r[4],
            "id": r[5], "id_div": r[6], "name": r[7], "status": r[8],
            "date": r[9], "wo_no": r[10], "equipment": r[11], "cat": r[12],
            "total_amt": r[13], "inv_no": r[14], "qb_id": r[15],
            "ref_no": r[16], "mbl_booking": r[17], "inv_amt": r[18],
            "paid_received": r[19], "qb_date": r[20] if len(r) > 20 else None,
        })
    return out


# ---------------------------------------------------------------------------
# Build engine
# ---------------------------------------------------------------------------

def build_today(spec: BuildSpec, input_workbook: str, age_delta: int) -> dict:
    yest_ar, yest_date = load_ar_sheet(input_workbook)
    qbo_col = load_qbo_collection(spec.qbo_collection)
    qbo_sch = load_qbo_schedule(spec.qbo_schedule)
    tab = load_tab_bank(spec.tab_bank)
    tms = load_tms_reconcile(spec.tms_reconcile)

    # Phase 1 — clone yesterday's AR
    today_ar: dict[str, ARRow] = {}
    for row in yest_ar:
        if row.inv:
            today_ar[row.inv] = replace(row)

    yest_amounts = {row.inv: row.amount for row in yest_ar if row.inv and row.amount is not None}

    # Phase 2 — apply collections
    collection_by_invoice: dict[str, list[dict]] = defaultdict(list)
    for line in qbo_col:
        if line["txn_type"] != "Invoice":
            continue
        collection_by_invoice[line["invoice_or_ref"]].append(line)

    today_str = spec.target_date.strftime("%m/%d/%Y")

    for inv, lines in collection_by_invoice.items():
        if inv not in today_ar:
            continue
        row = today_ar[inv]
        total_applied = sum((l["amount"] or 0) for l in lines)
        new_paid = (row.paid or 0) + total_applied
        new_balance = (row.amount or 0) - new_paid
        if abs(new_balance) < 0.01 or new_balance < -0.01:
            del today_ar[inv]
            continue
        check_no = lines[0]["txn_number"]
        row.paid = new_paid
        row.balance = new_balance
        if not row.memo:
            row.memo = f"Short paid {today_str} #{check_no}"

    # Phase 3 — add new invoices
    for sch in qbo_sch:
        inv = sch["inv"]
        if not inv or inv in today_ar:
            continue
        cust_field = sch["customer"] or ""
        cust_id = None
        cust_name = cust_field
        if cust_field.startswith("["):
            close = cust_field.find("]")
            if close > 0:
                cust_id = cust_field[1:close]
                cust_name = cust_field[close + 1:].strip()
        today_ar[inv] = ARRow(
            new_id=cust_id, company=cust_name, inv=inv,
            equipment=sch["cntr_chassis"], date=sch["date"],
            aging=0, ref_no=sch["ref"], mbl_no=sch["bl"],
            amount=sch["amount"], paid=0, balance=sch["amount"],
            ar_status=aging_bucket(0),
        )

    # Phase 4 — age every row by age_delta (calendar days since yesterday's workbook)
    for row in today_ar.values():
        if row.aging is not None:
            row.aging += age_delta
            row.ar_status = aging_bucket(row.aging)

    # Phase 5 — TMS reconcile
    tms_sheet_rows = []
    adjustment_rows = []
    new_invs = {sch["inv"] for sch in qbo_sch if sch["inv"]}

    for t in tms:
        if t["type"] != "AR":
            continue
        inv = t["inv_no"]
        if not inv:
            continue
        if inv in new_invs:
            if inv in today_ar:
                today_ar[inv].amount = t["total_amt"]
                today_ar[inv].balance = (t["total_amt"] or 0) - (today_ar[inv].paid or 0)
                today_ar[inv].wo = t["wo_no"]
            tms_sheet_rows.append(t)
        else:
            delta = t["total_amt"]
            if delta is not None and abs(delta) > 0.01:
                if inv in today_ar:
                    today_ar[inv].amount = t["inv_amt"]
                    today_ar[inv].balance = (t["inv_amt"] or 0) - (today_ar[inv].paid or 0)
                adjustment_rows.append({**t, "amount_difference": delta})
            else:
                tms_sheet_rows.append(t)

    return {
        "today_ar": list(today_ar.values()),
        "yest_date": yest_date,
        "tms_rows": tms_sheet_rows,
        "adjustment_rows": adjustment_rows,
        "new_invs": new_invs,
        "age_delta": age_delta,
    }


# ---------------------------------------------------------------------------
# Comparison
# ---------------------------------------------------------------------------

def _values_differ(a, b) -> bool:
    if a is None and b is None:
        return False
    if isinstance(a, (int, float)) and isinstance(b, (int, float)):
        return abs(a - b) > 0.01
    if a is None or b is None:
        if (isinstance(a, str) and not a.strip()) or (isinstance(b, str) and not b.strip()):
            return False
        return True
    return a != b


def compare_ar(produced: list[ARRow], target_path: str) -> dict:
    target_ar, _ = load_ar_sheet(target_path)
    prod_by_inv = {r.inv: r for r in produced if r.inv}
    tgt_by_inv = {r.inv: r for r in target_ar if r.inv}
    prod_invs = set(prod_by_inv)
    tgt_invs = set(tgt_by_inv)
    only_prod = prod_invs - tgt_invs
    only_tgt = tgt_invs - prod_invs
    common = prod_invs & tgt_invs

    field_diffs = defaultdict(list)
    matched_perfect = 0
    matched_with_diffs = 0
    for inv in common:
        p, t = prod_by_inv[inv], tgt_by_inv[inv]
        diffs = []
        for field_name in ("amount", "paid", "balance", "aging", "memo", "ar_status"):
            pv, tv = getattr(p, field_name), getattr(t, field_name)
            if _values_differ(pv, tv):
                diffs.append((field_name, pv, tv))
        if diffs:
            matched_with_diffs += 1
            for f, pv, tv in diffs:
                field_diffs[f].append((inv, pv, tv))
        else:
            matched_perfect += 1

    pct = (matched_perfect / max(1, len(target_ar))) * 100
    return {
        "produced_count": len(produced),
        "target_count": len(target_ar),
        "only_produced": sorted(only_prod),
        "only_target": sorted(only_tgt),
        "matched_perfect": matched_perfect,
        "matched_with_diffs": matched_with_diffs,
        "field_diffs": dict(field_diffs),
        "match_pct": pct,
    }


def report_summary(spec: BuildSpec, comp: dict, produced: dict, age_delta: int) -> None:
    print(f"\n{'='*78}")
    print(f"  Build {spec.target_date.strftime('%Y-%m-%d (%a)')} -> "
          f"target {os.path.basename(spec.target_workbook)}")
    print(f"  Input chained from yesterday's workbook | aging delta = {age_delta} day(s)")
    print(f"{'='*78}")
    print(f"  Match: {comp['match_pct']:.2f}%  ({comp['matched_perfect']} / {comp['target_count']} perfect)")
    print(f"  Produced rows: {comp['produced_count']}  |  Target rows: {comp['target_count']}")
    print(f"  Diffs:")
    print(f"    Field-level diffs: {comp['matched_with_diffs']} rows  |  "
          f"by field: { {f: len(v) for f, v in comp['field_diffs'].items()} }")
    print(f"    Only in produced (we have, target lacks): {len(comp['only_produced'])}")
    print(f"    Only in target (target has, we don't):    {len(comp['only_target'])}")

    if comp["field_diffs"]:
        for f, diffs in comp["field_diffs"].items():
            if f in ("amount", "balance") and diffs:
                print(f"    {f} sample diffs:")
                for inv, pv, tv in diffs[:3]:
                    print(f"      {inv}: produced={pv!r}  target={tv!r}")

    if comp["only_target"]:
        print(f"    Sample missing (only in target):")
        for inv in comp["only_target"][:5]:
            print(f"      {inv}")

    print(f"  TMS sheet rows: {len(produced['tms_rows'])}  |  ADJUSTMENT rows: {len(produced['adjustment_rows'])}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    builds = discover_builds()
    print(f"Discovered {len(builds)} builds:")
    for b in builds:
        print(f"  {b.target_date.date()}  <- {os.path.basename(b.target_workbook)}  ({os.path.relpath(b.folder, ROOT)})")
    print()

    overall = []

    for i, spec in enumerate(builds):
        input_wb = find_input_workbook(spec.target_date, builds)
        if not input_wb:
            print(f"!! Skipping {spec.target_date.date()} — can't find an input (prior) workbook")
            continue
        input_date = _parse_target_date(input_wb)
        age_delta = (spec.target_date - input_date).days if input_date else 1
        if age_delta < 1:
            age_delta = 1

        # Aging delta = calendar days between consecutive BUILD days (not workbook dates).
        # Build day = workbook date + 1 normally, but +3 if workbook date is a Friday
        # (she builds Friday's workbook on the following Monday).
        def build_day(d: datetime) -> datetime:
            return d + timedelta(days=3 if d.weekday() == 4 else 1)
        if input_date:
            age_delta = (build_day(spec.target_date) - build_day(input_date)).days
        if age_delta < 1:
            age_delta = 1

        produced = build_today(spec, input_wb, age_delta)
        comp = compare_ar(produced["today_ar"], spec.target_workbook)
        report_summary(spec, comp, produced, age_delta)
        overall.append((spec, comp))

    print()
    print("=" * 78)
    print(" SUMMARY ".center(78, "="))
    print("=" * 78)
    print(f"{'Date':<12}  {'Match %':>9}  {'Perfect':>8}  {'Diffs':>6}  {'OnlyProd':>9}  {'OnlyTgt':>8}")
    for spec, comp in overall:
        print(f"{spec.target_date.strftime('%Y-%m-%d'):<12}  "
              f"{comp['match_pct']:>8.2f}%  "
              f"{comp['matched_perfect']:>8}  "
              f"{comp['matched_with_diffs']:>6}  "
              f"{len(comp['only_produced']):>9}  "
              f"{len(comp['only_target']):>8}")


if __name__ == "__main__":
    main()
