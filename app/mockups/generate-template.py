"""Generate the sample customer-import template that demonstrates the proposed Fix 4 format.

Excel-only — the import will not accept CSV or JSON. Keeping it to one format
means everyone gets the dropdowns, tooltips, and Cheat Sheet.

Run: python app/mockups/generate-template.py
Output: app/mockups/customer-import-template.xlsx
"""
from pathlib import Path

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

OUT = Path(__file__).parent / "customer-import-template.xlsx"

# ── Data the template should demonstrate ──────────────────────────────
HEADERS = [
    "Code", "Name", "Email", "CC", "BCC",
    "Send Method", "Required Docs", "Notes",
]

# Real-looking sample rows. Mix of all three send methods and varied doc rules.
SAMPLE_ROWS = [
    ["APEXMA01", "APEX MARITIME CO",
     "ar@apexmaritime.com", "ar@ngltrans.net", "",
     "Email", "POD", ""],

    ["GLBCAR01", "GLOBAL CARGO LTD",
     "billing@globalcargo.com", "", "",
     "Email", "POD, BOL", ""],

    ["OECNYC01", "OEC GROUP — NEW YORK",
     "accounting@oecgroup.com", "ar@ngltrans.net", "",
     "OEC Two-Email", "", "POD goes via Gmail to DO Sender automatically"],

    ["TRURET01", "TRUE VALUE RETAIL SUPPORT CENTER",
     "", "", "",
     "Portal Upload", "", "Invoice + POD merged & uploaded to TranzAct"],

    ["SEAEXP01", "SEAWAY EXPRESS",
     "ap@seawayexpress.com, kim@seawayexpress.com", "", "",
     "Email", "", "Blank Required Docs = send any attachments on the invoice"],
]

VALID_SEND_METHODS = ["Email", "OEC Two-Email", "Portal Upload"]
VALID_DOCS = ["POD", "POL", "BOL", "PL", "DO"]

# ── Build workbook ──────────────────────────────────────────────────────
wb = Workbook()
ws = wb.active
ws.title = "Customers"

NGL_ORANGE = "EA580C"
LIGHT_ORANGE = "FFF7ED"
GRAY_BORDER = "E2E8F0"
HEADER_TEXT = "FFFFFF"

header_font = Font(name="Calibri", size=11, bold=True, color=HEADER_TEXT)
header_fill = PatternFill("solid", fgColor=NGL_ORANGE)
header_align = Alignment(horizontal="left", vertical="center", indent=1)

body_font = Font(name="Calibri", size=11)
note_font = Font(name="Calibri", size=10, italic=True, color="64748B")

thin = Side(style="thin", color=GRAY_BORDER)
cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)

# ── Header row ──
for col_idx, header in enumerate(HEADERS, start=1):
    cell = ws.cell(row=1, column=col_idx, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = cell_border

# Header comments — column-level help, hover to see
ws["A1"].comment = Comment(
    "REQUIRED. A short unique identifier for this customer.\n"
    "Convention: short name + 2 digits (e.g. APEXMA01, OECNYC01).",
    "NGL Template",
)
ws["B1"].comment = Comment(
    "REQUIRED. Customer name as it appears on QuickBooks invoices.",
    "NGL Template",
)
ws["C1"].comment = Comment(
    "Optional. Where invoice emails go. Multiple emails OK — "
    "separate with commas (e.g. ap@x.com, billing@x.com).",
    "NGL Template",
)
ws["F1"].comment = Comment(
    "REQUIRED. How invoices are delivered to this customer. Pick ONE from the dropdown:\n\n"
    "EMAIL  (most common)\n"
    "   → QuickBooks emails the invoice + every attachment together in one email.\n"
    "   → Use this for most customers.\n\n"
    "OEC TWO-EMAIL\n"
    "   → QuickBooks emails just the invoice (no POD attached).\n"
    "   → Then a SECOND email with just the POD goes via Gmail to the DO Sender.\n"
    "   → Use this only for OEC Group customers.\n\n"
    "PORTAL UPLOAD\n"
    "   → Invoice + POD get merged into one PDF.\n"
    "   → That merged PDF gets uploaded to the customer's TranzAct portal.\n"
    "   → No email is sent. Use this for portal-only customers like True Value.\n\n"
    "See the “Cheat Sheet” tab for more detail.",
    "NGL Template",
)
ws["G1"].comment = Comment(
    "Optional. Which supporting documents must be attached in QBO before we'll send the invoice.\n\n"
    "Valid values (type one or more, separated by commas):\n\n"
    "POD  = Proof of Delivery (the signed delivery receipt)\n"
    "POL  = Port of Loading document\n"
    "BOL  = Bill of Lading (the shipping contract)\n"
    "PL   = Packing List (what's inside the container)\n"
    "DO   = Delivery Order\n\n"
    "Examples:\n"
    "   POD\n"
    "   POD, BOL\n"
    "   POL, BOL, PL\n\n"
    "Leave BLANK to send whatever attachments happen to be on the QBO invoice.",
    "NGL Template",
)

# ── Sample rows ──
for row_offset, row_data in enumerate(SAMPLE_ROWS, start=2):
    for col_idx, value in enumerate(row_data, start=1):
        cell = ws.cell(row=row_offset, column=col_idx, value=value)
        cell.font = body_font
        cell.border = cell_border
        cell.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
        if col_idx == len(HEADERS):  # Notes column
            cell.font = note_font

# ── Column widths ──
COL_WIDTHS = {
    "A": 14, "B": 36, "C": 32, "D": 24, "E": 24,
    "F": 18, "G": 18, "H": 48,
}
for col, width in COL_WIDTHS.items():
    ws.column_dimensions[col].width = width

ws.row_dimensions[1].height = 28
ws.freeze_panes = "A2"  # freeze header

# ── Data validation: Send Method dropdown ──
sm_validation = DataValidation(
    type="list",
    formula1='"' + ",".join(VALID_SEND_METHODS) + '"',
    allow_blank=False,
    showErrorMessage=True,
    errorTitle="Not a valid Send Method",
    error="Pick one of: Email, OEC Two-Email, or Portal Upload.",
    promptTitle="Send Method",
    prompt="Click the dropdown arrow and pick one.",
    showInputMessage=True,
)
ws.add_data_validation(sm_validation)
sm_validation.add(f"F2:F1000")

# ── Banner row at the top of the sheet (above headers): nope, keep sheet clean.
# Instead we add a CHEAT SHEET tab with all the rules.

# ────────────────────────────────────────────────────────────────────────
# Sheet 2 — Cheat Sheet with all valid values + tips
# ────────────────────────────────────────────────────────────────────────
cs = wb.create_sheet("Cheat Sheet")

big_title_font = Font(name="Calibri", size=18, bold=True, color="0F172A")
intro_font = Font(name="Calibri", size=11, italic=True, color="64748B")
section_font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
section_fill = PatternFill("solid", fgColor=NGL_ORANGE)
option_name_font = Font(name="Calibri", size=12, bold=True, color="9A3412")
option_card_fill = PatternFill("solid", fgColor=LIGHT_ORANGE)
sub_label_font = Font(name="Calibri", size=10, bold=True, color="64748B")
value_font = Font(name="Calibri", size=11, color="334155")
red_warn_font = Font(name="Calibri", size=11, color="B91C1C", italic=True, bold=True)
hint_font = Font(name="Calibri", size=10, italic=True, color="64748B")

wrap_align = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
center_align = Alignment(horizontal="left", vertical="center", indent=1)

# ── Title ──
cs["A1"] = "How to fill out the Customers sheet"
cs["A1"].font = big_title_font
cs.row_dimensions[1].height = 28
cs["A2"] = "Open this tab any time you're not sure what to type. Hover the column headers on the Customers tab too — every header has a tooltip."
cs["A2"].font = intro_font
cs.row_dimensions[2].height = 20

# ── Helper to draw a section banner row ──
def section_banner(row, text):
    cs.merge_cells(start_row=row, end_row=row, start_column=1, end_column=4)
    cell = cs.cell(row=row, column=1, value=" " + text)
    cell.font = section_font
    cell.fill = section_fill
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    cs.row_dimensions[row].height = 32

# ── Helper to draw an "option card" (3 rows: name + sub-rows) ──
def option_card(start_row, name, who, what_happens, where):
    # Name row
    name_cell = cs.cell(row=start_row, column=1, value="  " + name)
    name_cell.font = option_name_font
    name_cell.fill = option_card_fill
    name_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    cs.merge_cells(start_row=start_row, end_row=start_row, start_column=1, end_column=4)
    cs.row_dimensions[start_row].height = 32

    # Sub-rows — bigger height so 2-3 lines of wrapped text fit comfortably
    rows = [
        ("Use when",       who,           42),
        ("What happens",   what_happens,  62),
        ("Where it goes",  where,         52),
    ]
    for i, (label, val, height) in enumerate(rows):
        r = start_row + 1 + i
        label_cell = cs.cell(row=r, column=1, value=label)
        label_cell.font = sub_label_font
        label_cell.alignment = Alignment(horizontal="right", vertical="top", indent=1)
        cs.merge_cells(start_row=r, end_row=r, start_column=2, end_column=4)
        val_cell = cs.cell(row=r, column=2, value=val)
        val_cell.font = value_font
        val_cell.alignment = wrap_align
        cs.row_dimensions[r].height = height
    # Spacer row after each card
    spacer = start_row + 4
    cs.row_dimensions[spacer].height = 8
    return start_row + 5

# ── REQUIRED COLUMNS section ──
section_banner(4, "REQUIRED COLUMNS — every row needs these")

required_rows = [
    ("Code", "A short unique ID like APEXMA01. No spaces. Every row must have one.", "APEXMA01"),
    ("Name", "Customer name as it appears on QuickBooks invoices.", "APEX MARITIME CO"),
    ("Send Method", "How this customer's invoices get delivered. Pick from the dropdown.", "Email"),
]
r = 5
# Header row for the table
cs.cell(row=r, column=1, value="Column").font = sub_label_font
cs.cell(row=r, column=1).alignment = Alignment(horizontal="left", vertical="center", indent=1)
cs.cell(row=r, column=2, value="What it is").font = sub_label_font
cs.cell(row=r, column=2).alignment = Alignment(horizontal="left", vertical="center", indent=1)
cs.merge_cells(start_row=r, end_row=r, start_column=2, end_column=3)
cs.cell(row=r, column=4, value="Example").font = sub_label_font
cs.cell(row=r, column=4).alignment = Alignment(horizontal="left", vertical="center", indent=1)
cs.row_dimensions[r].height = 22
r += 1
for col, desc, ex in required_rows:
    cs.cell(row=r, column=1, value=col).font = Font(name="Calibri", size=11, bold=True, color="0F172A")
    cs.cell(row=r, column=1).alignment = Alignment(horizontal="left", vertical="center", indent=1)
    desc_cell = cs.cell(row=r, column=2, value=desc)
    desc_cell.font = value_font
    desc_cell.alignment = wrap_align
    cs.merge_cells(start_row=r, end_row=r, start_column=2, end_column=3)
    ex_cell = cs.cell(row=r, column=4, value=ex)
    ex_cell.font = Font(name="Consolas", size=10, color="334155")
    ex_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    cs.row_dimensions[r].height = 38
    r += 1

# ── SEND METHOD section ──
r += 1  # blank row
section_banner(r, "SEND METHOD — pick exactly ONE per customer")
r += 1

r = option_card(
    r,
    "EMAIL  (use this for most customers)",
    "Most customers — anyone who just gets a normal invoice email from QuickBooks.",
    "QuickBooks sends ONE email containing the invoice PDF + every attachment that's on the invoice (POD, BOL, etc).",
    "To the email addresses in the Email / CC / BCC columns.",
)

r = option_card(
    r,
    "OEC TWO-EMAIL  (only for OEC Group customers)",
    "Customers in the OEC Group family. They want the invoice and the POD in two separate emails — not one combined.",
    "Email 1: QuickBooks sends the invoice (no POD attached).\nEmail 2: A second email goes out via Gmail with just the POD.",
    "Email 1 → the addresses in the Email column.\nEmail 2 → the DO Sender's email (pulled automatically from the invoice's container info).",
)

r = option_card(
    r,
    "PORTAL UPLOAD  (only for TranzAct portal customers)",
    "Customers like True Value Retail who don't want emails — they want their docs uploaded to the TranzAct portal.",
    "The agent merges the invoice PDF + POD into one combined PDF, then opens the customer's TranzAct portal and uploads it.",
    "Customer's TranzAct portal — no email is sent to anyone.",
)

# ── REQUIRED DOCS section ──
r += 1
section_banner(r, "REQUIRED DOCS — what supporting documents must be attached")
r += 1

cs.merge_cells(start_row=r, end_row=r, start_column=1, end_column=4)
intro = cs.cell(row=r, column=1, value=(
    "  This column is OPTIONAL. If you list one or more docs here, the invoice will only get sent "
    "once those docs are attached in QBO. Leave it BLANK to send whatever attachments happen to be on the invoice."
))
intro.font = value_font
intro.alignment = wrap_align
cs.row_dimensions[r].height = 44
r += 2

# Header row
cs.cell(row=r, column=1, value="Type this").font = sub_label_font
cs.cell(row=r, column=1).alignment = Alignment(horizontal="left", vertical="center", indent=1)
cs.cell(row=r, column=2, value="Stands for").font = sub_label_font
cs.cell(row=r, column=2).alignment = Alignment(horizontal="left", vertical="center", indent=1)
cs.cell(row=r, column=3, value="What the document is").font = sub_label_font
cs.cell(row=r, column=3).alignment = Alignment(horizontal="left", vertical="center", indent=1)
cs.merge_cells(start_row=r, end_row=r, start_column=3, end_column=4)
cs.row_dimensions[r].height = 22
r += 1

doc_rows = [
    ("POD", "Proof of Delivery",   "The signed delivery receipt — proves the freight was delivered to the consignee."),
    ("POL", "Port of Loading",     "Document showing the port where the container was loaded onto the ship."),
    ("BOL", "Bill of Lading",      "The shipping contract between the shipper and the carrier — basically the freight's title document."),
    ("PL",  "Packing List",        "An itemized list of what's inside the container."),
    ("DO",  "Delivery Order",      "The instruction from the carrier authorizing release of the freight to the consignee."),
]
for typ, full, desc in doc_rows:
    code_cell = cs.cell(row=r, column=1, value=typ)
    code_cell.font = Font(name="Consolas", size=12, bold=True, color="EA580C")
    code_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    full_cell = cs.cell(row=r, column=2, value=full)
    full_cell.font = Font(name="Calibri", size=11, bold=True, color="0F172A")
    full_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    desc_cell = cs.cell(row=r, column=3, value=desc)
    desc_cell.font = value_font
    cs.merge_cells(start_row=r, end_row=r, start_column=3, end_column=4)
    desc_cell.alignment = wrap_align
    cs.row_dimensions[r].height = 44
    r += 1

# Format examples
r += 1
cs.cell(row=r, column=1, value="Format examples").font = sub_label_font
cs.cell(row=r, column=1).alignment = Alignment(horizontal="left", vertical="center", indent=1)
cs.row_dimensions[r].height = 22
r += 1
fmt_examples = [
    ("POD",            "Just require the POD."),
    ("POD, BOL",       "Require both POD and BOL."),
    ("POL, BOL, PL",   "Require all three."),
    ("(blank)",        "Send whatever's on the invoice — no specific requirement."),
]
for typed, meaning in fmt_examples:
    typed_cell = cs.cell(row=r, column=1, value=typed)
    typed_cell.font = Font(name="Consolas", size=11, color="334155")
    typed_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    cs.merge_cells(start_row=r, end_row=r, start_column=2, end_column=4)
    meaning_cell = cs.cell(row=r, column=2, value="→ " + meaning)
    meaning_cell.font = value_font
    meaning_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    cs.row_dimensions[r].height = 26
    r += 1

# ── EMAIL COLUMNS section ──
r += 2
section_banner(r, "EMAIL / CC / BCC — recipient addresses")
r += 1
cs.merge_cells(start_row=r, end_row=r, start_column=1, end_column=4)
em = cs.cell(row=r, column=1, value=(
    "  Type one email per customer, OR several separated by commas in the same cell.\n"
    "  Example:   ap@acme.com, billing@acme.com, kim@acme.com"
))
em.font = value_font
em.alignment = wrap_align
cs.row_dimensions[r].height = 52

# ── TIPS section ──
r += 2
section_banner(r, "TIPS")
r += 1
tips = [
    "When you save the file, Excel may ask about format — say YES to keep .xlsx.",
    "If you're editing an exported file from the Customers tool, you can re-import it directly — the format already matches.",
    "Customer codes are case-insensitive on import (apexma01 = APEXMA01).",
    "If a customer code already exists, you'll see a warning before anything is overwritten.",
]
for tip in tips:
    cs.merge_cells(start_row=r, end_row=r, start_column=1, end_column=4)
    cell = cs.cell(row=r, column=1, value="  •  " + tip)
    cell.font = hint_font
    cell.alignment = wrap_align
    cs.row_dimensions[r].height = 28
    r += 1

# Cheat sheet column widths — generous so wrapped text breathes on a fullscreen monitor
cs.column_dimensions["A"].width = 22
cs.column_dimensions["B"].width = 32
cs.column_dimensions["C"].width = 60
cs.column_dimensions["D"].width = 22
cs.sheet_view.showGridLines = False

# ── Save xlsx ──
wb.save(OUT)
print(f"Wrote {OUT}")
print(f"  - Sheet 1 (Customers): {len(SAMPLE_ROWS)} sample rows + dropdown on Send Method column")
print(f"  - Sheet 2 (Cheat Sheet): valid values + tips")
